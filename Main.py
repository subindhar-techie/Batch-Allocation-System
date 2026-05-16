import os
import glob
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    import sys
    import warnings
    warnings.warn("openpyxl is not installed. Please run `pip install openpyxl`.")
    messagebox = tk.messagebox if 'tkinter' in sys.modules else None
    
# Ensure long path format for Windows path limits (> 260 character support)
def normalize_long_path(path):
    abs_path = os.path.abspath(path)
    if os.name == 'nt' and not abs_path.startswith('\\\\?\\'):
        abs_path = '\\\\?\\' + abs_path
    return abs_path

def parse_file(file_path):
    file_path = normalize_long_path(file_path)
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
    except UnicodeDecodeError:
        try:
            with open(file_path, 'r', encoding='ISO-8859-1') as f:
                lines = f.readlines()
        except Exception as e:
            return None, str(e)
    except Exception as e:
        return None, str(e)
            
    headers = {}
    data_lines = []
    
    # Parse headers (first 15 lines max)
    for i in range(min(15, len(lines))):
        line = lines[i].strip()
        if ':' in line:
            parts = line.split(':', 1)
            key = parts[0].strip()
            val = parts[1].strip()
            headers[key] = val
            
    # Typically headers are up to line 13-14, line 15 is column names.
    # Data starts from line 16 (index 15) as per requirements.
    for i in range(15, len(lines)):
        line = lines[i].strip()
        if line:
            # Safely handle tab-separated, pipe-separated, or space-separated files
            if '\t' in line:
                row = line.split('\t')
            elif '|' in line:
                row = [x.strip() for x in line.split('|')]
            else:
                row = line.split()
            
            data_lines.append(row)
            
    if not data_lines:
        return None, "No data found starting line 16"
        
    first_row = data_lines[0]
    last_row = data_lines[-1]
    
    # Extract headers according to input file spec
    po_number = headers.get('PO Number', '')
    batch_no = headers.get('Batch NO', '')
    circle = headers.get('Circle', '')
    qty = headers.get('SIM Quantity', '')
    
    # Assuming standard columns: IMPU | IMPI | IMSI | IMSII | ICCID | PIN1 ...
    # Wait, user prompt lists columns: 
    # IMPU | IMPI | IMSI | IMSII | ICCID | PIN1 | PUK1 | PIN2 | PUK2 | EncryptedKi | EncryptedOPC | TranskeyIndex | OPKeyIndex | MSN | EID
    # So IMSI is at index 2, and ICCID is at index 4
    try:
        start_imsi = first_row[2]
        start_iccid = first_row[4]
        end_imsi = last_row[2]
        end_iccid = last_row[4]
    except IndexError:
        return None, f"Data row format issue. Not enough columns. Found {len(first_row)} columns."
        
    return {
        'PO NUMBER': po_number,
        'CIRCLE': circle,
        'BATCH': batch_no,
        'QTY': qty,
        'START IMSI': start_imsi,
        'START ICCID': start_iccid,
        'END IMSI': end_imsi,
        'END ICCID': end_iccid
    }, None

def process_folder(folder_path, status_label, progress_bar):
    folder_path = normalize_long_path(folder_path)
    
    # Find all text files in all subdirectories
    txt_files = []
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.lower().endswith('.txt'):
                txt_files.append(os.path.join(root, file))
                
    if not txt_files:
        messagebox.showinfo("Info", "No .txt files found in the selected folder.")
        return
        
    status_label.config(text=f"Found {len(txt_files)} files. Processing...")
    progress_bar['maximum'] = len(txt_files)
    progress_bar['value'] = 0
    status_label.update()
    
    results = []
    errors = []
    
    for i, fpath in enumerate(txt_files):
        data, err = parse_file(fpath)
        if data:
            data['FILENAME'] = os.path.basename(fpath)
            results.append(data)
        else:
            errors.append(f"File {os.path.basename(fpath)}: {err}")
            
        progress_bar['value'] = i + 1
        status_label.update()
        
    if not results:
        messagebox.showerror("Error", "Could not parse any valid data from the files.")
        if errors:
            print("\n".join(errors))
        status_label.config(text="Failed.")
        return
        
    # Determine expected PO, Circle, and Pattern from the first valid filename
    po_for_name = None
    circle_for_name = None
    expected_pattern = None
    expected_pattern_parts = None
    for res in results:
        fname = res.get('FILENAME', '')
        parts = os.path.splitext(fname)[0].split('_')
        if len(parts) >= 5:
            po_for_name = parts[1]
            circle_for_name = parts[4]
            expected_pattern_parts = list(parts)
            expected_pattern_parts[3] = "<BATCH>"
            expected_pattern_parts[-1] = "<TIMESTAMP>"
            expected_pattern = "_".join(expected_pattern_parts)
            break
            
    if not po_for_name:
        po_for_name = results[0]['PO NUMBER']
        circle_for_name = results[0]['CIRCLE']
    
    filename_errors = {} # filename -> list of error strings
    batch_files = {} # batch -> list of filenames
    duplicate_batches = set()
    
    # Phase 1: Filename validation
    for res in results:
        filename = res.get('FILENAME', 'Unknown File')
        base_name = os.path.splitext(filename)[0]
        parts = base_name.split('_')
        
        if len(parts) >= 5:
            file_po = parts[1]
            file_batch = parts[3]
            file_circle = parts[4]
            
            if file_po != po_for_name:
                filename_errors.setdefault(filename, []).append(f"Different PO found in filename: {file_po} (Expected: {po_for_name})")
            if file_circle != circle_for_name:
                filename_errors.setdefault(filename, []).append(f"Different Circle found in filename: {file_circle} (Expected: {circle_for_name})")
                
            if expected_pattern_parts:
                current_pattern_parts = list(parts)
                current_pattern_parts[3] = "<BATCH>"
                current_pattern_parts[-1] = "<TIMESTAMP>"
                
                if current_pattern_parts != expected_pattern_parts:
                    if len(current_pattern_parts) != len(expected_pattern_parts):
                        filename_errors.setdefault(filename, []).append(f"Filename has incorrect structure (Found {len(parts)} segments, expected {len(expected_pattern_parts)})")
                    else:
                        mismatches = []
                        for idx, (curr, exp) in enumerate(zip(current_pattern_parts, expected_pattern_parts)):
                            if curr != exp and exp not in ("<BATCH>", "<TIMESTAMP>"):
                                mismatches.append(f"'{curr}' instead of '{exp}'")
                        
                        if mismatches:
                            filename_errors.setdefault(filename, []).append(f"Filename pattern mismatch: Found {', '.join(mismatches)}")
                        else:
                            display_pattern = expected_pattern.replace('<BATCH>', '*').replace('<TIMESTAMP>', '*')
                            filename_errors.setdefault(filename, []).append(f"Filename pattern mismatch (Expected format: {display_pattern})")
                    
            batch = file_batch
        else:
            if filename != 'Unknown File':
                filename_errors.setdefault(filename, []).append("Filename format invalid for metadata extraction")
            batch = res.get('BATCH', 'Unknown')
            
        if batch in batch_files:
            batch_files[batch].append(filename)
            duplicate_batches.add(batch)
        else:
            batch_files[batch] = [filename]
            
    # Add duplicate batch errors
    general_errors = []
    if duplicate_batches:
        for b in duplicate_batches:
            general_errors.append(f"Duplication found in this batch: {b}.")
                
    if filename_errors or general_errors:
        error_lines = general_errors.copy()
        for fname, errs in filename_errors.items():
            error_lines.append(f"{fname} - {', '.join(errs)}")
            
        error_msg = "Filename Validation failed with the following errors:\n\n" + "\n".join(error_lines[:10])
        if len(error_lines) > 10:
            error_msg += f"\n...and {len(error_lines) - 10} more files."
            
        messagebox.showerror("Validation Error", error_msg)
        status_label.config(text="Validation failed.")
        
        if 'error_files_var' in globals():
            ui_msg = "Filename Validation Errors:\n" + "\n".join(error_lines[:15])
            if len(error_lines) > 15:
                ui_msg += f"\n...and {len(error_lines) - 15} more files."
            error_files_var.set(ui_msg)
        return
        
    # Phase 2: Internal Header vs Filename validation
    internal_errors = {}
    for res in results:
        filename = res.get('FILENAME', 'Unknown File')
        base_name = os.path.splitext(filename)[0]
        parts = base_name.split('_')
        
        if len(parts) >= 5:
            file_po = parts[1]
            file_batch = parts[3]
            file_circle = parts[4]
            
            if res['PO NUMBER'] != file_po:
                internal_errors.setdefault(filename, []).append(f"PO inside file ({res['PO NUMBER']}) does not match filename PO ({file_po})")
            if res['BATCH'] != file_batch:
                internal_errors.setdefault(filename, []).append(f"Batch inside file ({res['BATCH']}) does not match filename Batch ({file_batch})")
            if res['CIRCLE'] != file_circle:
                internal_errors.setdefault(filename, []).append(f"Circle inside file ({res['CIRCLE']}) does not match filename Circle ({file_circle})")

    if internal_errors:
        error_lines = []
        for fname, errs in internal_errors.items():
            error_lines.append(f"{fname} - {', '.join(errs)}")
            
        error_msg = "Internal Content Validation failed with the following errors:\n\n" + "\n".join(error_lines[:10])
        if len(error_lines) > 10:
            error_msg += f"\n...and {len(error_lines) - 10} more files."
            
        messagebox.showerror("Validation Error", error_msg)
        status_label.config(text="Validation failed.")
        
        if 'error_files_var' in globals():
            ui_msg = "Internal Validation Errors:\n" + "\n".join(error_lines[:15])
            if len(error_lines) > 15:
                ui_msg += f"\n...and {len(error_lines) - 15} more files."
            error_files_var.set(ui_msg)
        return

    # If successful, clear the error files display
    if 'error_files_var' in globals():
        error_files_var.set("")

    # Sort results by BATCH in ascending order
    results.sort(key=lambda x: int(x['BATCH']) if str(x['BATCH']).isdigit() else x['BATCH'])
        
    try:
        total_qty = sum(int(res['QTY']) for res in results if str(res['QTY']).isdigit())
        
        # Create the workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Add header metadata
        ws.append(['OPERATOR NAME', 'RJIL'])
        ws.append(['PO NO', po_for_name])
        ws.append(['CIRCLE', circle_for_name])
        ws.append(['TOTAL QTY', total_qty])
        ws.append([]) # Empty line
        
        # Headers definitions
        headers = ['PO Number', 'Circle', 'Batch', 'Qty', 'Start IMSI', 'End IMSI', 'Start ICCID', 'End ICCID']
        ws.append(headers)
        
        # Append data setting it properly as text format
        for row_idx, result in enumerate(results, start=7):
            row_data = [
                result['PO NUMBER'],
                result['CIRCLE'],
                result['BATCH'],
                result['QTY'],
                result['START IMSI'],
                result['END IMSI'],
                result['START ICCID'],
                result['END ICCID']
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.number_format = '@' # Force Excel to treat as Text
                
        # Set styling specs from requirements
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid') # Dark Blue
        data_font = Font(name='Calibri', size=11)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        # Define bold inner and outer borders
        bold_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        
        # Apply styles to Top Metadata (Rows 1-4)
        for row_idx in range(1, 5):
            # Column A
            cell_a = ws.cell(row=row_idx, column=1)
            cell_a.font = header_font
            cell_a.fill = header_fill
            cell_a.border = bold_border
            
            # Column B
            cell_b = ws.cell(row=row_idx, column=2)
            cell_b.font = Font(name='Calibri', size=11, bold=True)
            cell_b.alignment = left_align
            cell_b.border = bold_border
            cell_b.number_format = '@'
            
        # Apply styles to Data Headers
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=6, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = bold_border
            
            # Slightly widen columns for better aesthetics
            ws.column_dimensions[cell.column_letter].width = 22
            
        # Apply styles to Data
        for row_idx in range(7, len(results) + 7):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.alignment = center_align
                cell.border = bold_border
    
        # Save dynamically based on data content
        output_path = os.path.join(folder_path, f"{po_for_name}_{circle_for_name}.xlsx")
        
        wb.save(output_path)
        
        if 'po_summary_var' in globals():
            po_summary_var.set(f"PO Number : {po_for_name}")
            files_summary_var.set(f"Found Input Files : {len(txt_files)}")
            circle_summary_var.set(f"CIRCLE : {circle_for_name}")
            qty_summary_var.set(f"TOTAL QTY : {total_qty}")
        
        msg = f"Report successfully generated!\nSaved at:\n{output_path}"
        if errors:
            msg += f"\n\nNote: {len(errors)} files had errors and were skipped."
            print("Errors during parsing:")
            for e in errors:
                print(e)
        messagebox.showinfo("Success", msg)
        
    except Exception as e:
        messagebox.showerror("Error", f"Could not save Excel file:\n{str(e)}\n\nPlease ensure the file is not currently open.")

    status_label.config(text="Done.")


def select_folder():
    folder_selected = filedialog.askdirectory(title="Select PO Folder")
    if folder_selected:
        folder_path_var.set(folder_selected)

def run_process():
    # Make sure openpyxl is available
    if 'openpyxl' not in sys.modules:
        messagebox.showerror("Missing Dependency", "The required library 'openpyxl' is not installed.\nPlease run 'pip install openpyxl' to continue.")
        return
        
    folder = folder_path_var.get()
    if not folder:
        messagebox.showwarning("Warning", "Please select a folder first.")
        return
    if not os.path.isdir(normalize_long_path(folder)):
        messagebox.showwarning("Warning", "The selected path is not a valid folder.")
        return
        
    process_folder(folder, status_label, progress_bar)

def clear_all():
    folder_path_var.set("")
    status_label.config(text="Ready.")
    progress_bar['value'] = 0
    if 'po_summary_var' in globals():
        po_summary_var.set("PO Number : -")
        files_summary_var.set("Found Input Files : -")
        circle_summary_var.set("CIRCLE : -")
        qty_summary_var.set("TOTAL QTY : -")
    if 'error_files_var' in globals():
        error_files_var.set("")

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

if __name__ == "__main__":
    import sys
    root = tk.Tk()
    root.title("Batch Allocation System")
    root.geometry("520x480")
    root.resizable(False, False)
    
    logo_img = None
    try:
        img_path = resource_path("Reliance_Jio_Logo.png")
        if os.path.exists(img_path):
            logo_img = tk.PhotoImage(file=img_path)
            root.iconphoto(False, logo_img)
    except Exception as e:
        print("Could not load icon:", e)
    
    style = ttk.Style(root)
    if "clam" in style.theme_names():
        style.theme_use("clam")
        
    style.configure("DarkBlue.Horizontal.TProgressbar", background="#002060", troughcolor="white", thickness=15)
    
    frame = tk.Frame(root, padx=20, pady=20)
    frame.pack(expand=True, fill="both")
    
    tk.Label(frame, text="Select Input Folder:", font=("Arial", 10, "bold")).pack(anchor="w", pady=(0, 5))
    
    folder_path_var = tk.StringVar()
    path_frame = tk.Frame(frame)
    path_frame.pack(fill="x", pady=5)
    
    entry = tk.Entry(path_frame, textvariable=folder_path_var, font=("Arial", 10))
    entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
    
    btn_browse = tk.Button(path_frame, text="Browse", command=select_folder, width=10)
    btn_browse.pack(side="right")
    
    button_frame = tk.Frame(frame)
    button_frame.pack(fill="x", pady=10)
    
    btn_run = tk.Button(button_frame, text="Generate Excel Report", command=run_process, font=("Arial", 10, "bold"), bg="#4CAF50", fg="white", pady=5)
    btn_run.pack(side="left", fill="x", expand=True, padx=(0, 5))
    
    btn_clear = tk.Button(button_frame, text="Clear All", command=clear_all, font=("Arial", 10, "bold"), bg="#f44336", fg="white", pady=5)
    btn_clear.pack(side="right", fill="x", expand=True, padx=(5, 0))
    
    progress_bar = ttk.Progressbar(frame, orient="horizontal", mode="determinate", style="DarkBlue.Horizontal.TProgressbar")
    progress_bar.pack(fill="x", pady=5)
    
    status_label = tk.Label(frame, text="Ready.", font=("Arial", 9))
    status_label.pack(anchor="w", pady=(0, 5))
    
    summary_frame = tk.LabelFrame(frame, text="Overall Summary", font=("Arial", 10, "bold"), padx=10, pady=10)
    summary_frame.pack(fill="x", pady=5)
    
    po_summary_var = tk.StringVar(value="PO Number : -")
    files_summary_var = tk.StringVar(value="Found Files : -")
    circle_summary_var = tk.StringVar(value="CIRCLE : -")
    qty_summary_var = tk.StringVar(value="TOTAL QTY : -")
    
    tk.Label(summary_frame, textvariable=po_summary_var, font=("Arial", 9)).pack(anchor="w")
    tk.Label(summary_frame, textvariable=files_summary_var, font=("Arial", 9)).pack(anchor="w")
    tk.Label(summary_frame, textvariable=circle_summary_var, font=("Arial", 9)).pack(anchor="w")
    tk.Label(summary_frame, textvariable=qty_summary_var, font=("Arial", 9)).pack(anchor="w")
    
    error_files_var = tk.StringVar(value="")
    error_label = tk.Label(frame, textvariable=error_files_var, font=("Arial", 9), fg="red", justify="left", wraplength=480)
    error_label.pack(anchor="w", pady=(5, 0))
    
    root.mainloop()
